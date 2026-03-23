"""
İTÜ Sürdürülebilirlik Haberleri Scraper
========================================
Kaynak: https://haberler.itu.edu.tr/surdurulebilir

Kurulum:
    pip install requests beautifulsoup4 openpyxl

Çıktı:
    - itu_haberler.json
    - itu_haberler.xlsx
"""

import requests
from bs4 import BeautifulSoup
import json
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BASE_URL = "https://haberler.itu.edu.tr"
LIST_URL = f"{BASE_URL}/surdurulebilir"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


# ─── Scraping ────────────────────────────────────────────────────────────────

def get_article_links(list_url: str) -> list[dict]:
    response = requests.get(list_url, headers=HEADERS, timeout=15)
    response.raise_for_status()
    response.encoding = "utf-8"

    soup = BeautifulSoup(response.text, "html.parser")
    articles = []
    seen_urls = set()

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        if "/haberdetay/" not in href:
            continue

        full_url = href if href.startswith("http") else BASE_URL + href
        if full_url in seen_urls:
            continue
        seen_urls.add(full_url)

        title_tag = a_tag.find(["h6", "h5", "h4"])
        title = title_tag.get_text(strip=True) if title_tag else ""

        parts = href.strip("/").split("/")
        try:
            date_str = f"{parts[1]}-{parts[2]}-{parts[3]}"
        except IndexError:
            date_str = ""

        articles.append({"title": title, "url": full_url, "date": date_str})

    return articles


def get_article_detail(url: str) -> dict:
    """
    Haber detay sayfasından başlık ve içeriği çeker.
    Başlık her zaman <h1> den alınır — liste sayfasındaki HTML yapısından bağımsız.
    Bu sayede liste sayfasında bazı haberlerde başlık boş gelse de buradan doğru çekilir.
    """
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.raise_for_status()
        response.encoding = "utf-8"
    except requests.RequestException as e:
        return {"title": "", "content": f"[İçerik alınamadı: {e}]"}

    soup = BeautifulSoup(response.text, "html.parser")

    # Başlık: detay sayfasındaki <h1> — her haber için kesin ve doğru kaynak
    h1 = soup.find("h1")
    title = h1.get_text(strip=True) if h1 else ""

    # İçerik: h1 den sonra gelen paragraflar
    content_parts = []
    if h1:
        for sibling in h1.find_next_siblings():
            if sibling.name in ["footer", "nav", "h6"]:
                break
            text = sibling.get_text(separator=" ", strip=True)
            if text:
                content_parts.append(text)

    return {"title": title, "content": "\n\n".join(content_parts)}


def scrape(delay: float = 0.5) -> list[dict]:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Liste sayfası çekiliyor...")
    articles = get_article_links(LIST_URL)
    print(f"  → {len(articles)} haber bulundu.\n")

    results = []
    for i, article in enumerate(articles, 1):
        detail = get_article_detail(article["url"])
        # Başlık: detay sayfasından gelen <h1> kullan (liste sayfasındaki boşsa da dolu gelir)
        article["title"] = detail["title"]
        article["content"] = detail["content"]
        print(f"[{i:02d}/{len(articles)}] {article['date']} | {article['title'][:70]}")
        results.append(article)
        time.sleep(delay)

    return results


# ─── JSON çıktısı ────────────────────────────────────────────────────────────

def save_to_json(data: list[dict], filename: str = "itu_haberler.json") -> None:
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n✓ JSON kaydedildi: {filename}")


# ─── Excel çıktısı ───────────────────────────────────────────────────────────

# Renk sabitleri
COLOR_HEADER_BG   = "1A3A5C"   # Koyu lacivert — header arka planı
COLOR_HEADER_FONT = "FFFFFF"   # Beyaz — header yazısı
COLOR_ROW_ODD     = "EAF2FB"   # Açık mavi — tek satırlar
COLOR_ROW_EVEN    = "FFFFFF"   # Beyaz — çift satırlar
COLOR_BORDER      = "BDC3C7"   # Gri kenarlık
COLOR_TITLE_FONT  = "1A3A5C"   # Lacivert — başlık hücresi yazısı
COLOR_SUMMARY_BG  = "D6EAF8"   # Özet satırı arka planı


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def save_to_excel(data: list[dict], filename: str = "itu_haberler.xlsx") -> None:
    wb = Workbook()

    # ── Sayfa 1: Haber Listesi ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Haber Listesi"

    # Başlık satırı (A1 birleşik)
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "İTÜ Sürdürülebilirlik Haberleri"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=COLOR_TITLE_FONT)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", start_color="D6EAF8")
    ws.row_dimensions[1].height = 30

    # Oluşturma tarihi (A2)
    ws["A2"] = f"Oluşturulma tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="7F8C8D")
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 16

    # Boş satır
    ws.row_dimensions[3].height = 6

    # Sütun başlıkları (4. satır)
    col_headers = ["#", "Tarih", "Başlık", "URL"]
    header_row = 4
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_FONT)
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[header_row].height = 22

    # Haber satırları (5. satırdan itibaren)
    for i, article in enumerate(data):
        row = header_row + 1 + i
        bg = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
        fill = PatternFill("solid", start_color=bg)

        values = [i + 1, article.get("date", ""), article.get("title", ""), article.get("url", "")]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 3))

        # URL'yi tıklanabilir hyperlink yap
        url_cell = ws.cell(row=row, column=4)
        url_cell.hyperlink = article.get("url", "")
        url_cell.font = Font(name="Arial", size=10, color="2471A3", underline="single")

        ws.row_dimensions[row].height = 18

    # Özet satırı
    summary_row = header_row + 1 + len(data)
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    summary_cell = ws[f"A{summary_row}"]
    summary_cell.value = f"Toplam haber sayısı:"
    summary_cell.font = Font(name="Arial", bold=True, size=10)
    summary_cell.fill = PatternFill("solid", start_color=COLOR_SUMMARY_BG)
    summary_cell.border = _thin_border()
    summary_cell.alignment = Alignment(horizontal="right", vertical="center")

    count_cell = ws[f"C{summary_row}"]
    count_cell.value = f'=COUNTA(C5:C{summary_row - 1})'
    count_cell.font = Font(name="Arial", bold=True, size=10)
    count_cell.fill = PatternFill("solid", start_color=COLOR_SUMMARY_BG)
    count_cell.border = _thin_border()
    count_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[summary_row].height = 20

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 75

    # Başlık satırını dondur
    ws.freeze_panes = "A5"

    # Filtre ekle
    ws.auto_filter.ref = f"A{header_row}:D{header_row + len(data)}"

    # ── Sayfa 2: Haber İçerikleri ────────────────────────────────────────────
    ws2 = wb.create_sheet("Haber İçerikleri")

    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "İTÜ Sürdürülebilirlik Haberleri — İçerikler"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_TITLE_FONT)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2["A1"].fill = PatternFill("solid", start_color="D6EAF8")
    ws2.row_dimensions[1].height = 30

    ws2.row_dimensions[2].height = 6

    for col_idx, header in enumerate(["Tarih", "Başlık", "İçerik"], 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_FONT)
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws2.row_dimensions[3].height = 22

    for i, article in enumerate(data):
        row = 4 + i
        bg = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
        fill = PatternFill("solid", start_color=bg)

        for col_idx, value in enumerate(
            [article.get("date", ""), article.get("title", ""), article.get("content", "")], 1
        ):
            cell = ws2.cell(row=row, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws2.row_dimensions[row].height = 80

    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 90
    ws2.freeze_panes = "A4"

    wb.save(filename)
    print(f"✓ Excel kaydedildi: {filename}")


# ─── Ana akış ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    haberler = scrape(delay=0.5)

    save_to_json(haberler)
    save_to_excel(haberler)

    print(f"\n--- ÖZET ---")
    print(f"Toplam haber : {len(haberler)}")
    if haberler:
        print(f"Tarih aralığı: {haberler[-1]['date']} → {haberler[0]['date']}")