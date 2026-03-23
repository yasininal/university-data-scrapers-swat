"""
EngiRank Türk Üniversiteleri Scraper
======================================
Kaynak : https://engirank.eu/ranking/2025/
Çıktı  : engirank_turkiye_YYYYMMDD_HHMMSS.xlsx  +  .json

Kurulum:
    pip install requests beautifulsoup4 openpyxl

Çalıştırma:
    python engirank_scraper.py
"""

import requests
from bs4 import BeautifulSoup
import json
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Sabitler ────────────────────────────────────────────────────────────────

BASE_URL = "https://engirank.eu"

# Çekilecek tüm sıralama kategorileri (yıl / konu kodu)
SUBJECTS = {
    "Institutional"                             : "all",
    "Chemical Engineering"                      : "che",
    "Civil Engineering"                         : "civ",
    "Electrical, Electronic & Information Eng." : "eei",
    "Environmental Engineering"                 : "env",
    "Materials Engineering"                     : "mat",
    "Mechanical Engineering"                    : "mec",
    "Medical Engineering"                       : "med",
}

YEARS   = [2025, 2024, 2023]
COUNTRY = "Turkiye"          # Sayfadaki ülke adı (Türkçe karakter yok)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# ─── Excel renkleri ──────────────────────────────────────────────────────────

C_HDR_BG    = "1B3A5C"   # Koyu lacivert — başlık arka planı
C_HDR_FONT  = "FFFFFF"   # Beyaz — başlık yazısı
C_ODD       = "EAF2FB"   # Açık mavi — tek satır
C_EVEN      = "FFFFFF"   # Beyaz — çift satır
C_BORDER    = "BDC3C7"   # Gri kenarlık
C_TITLE_BG  = "D6EAF8"   # Başlık satırı arka planı
C_TITLE_FG  = "1B3A5C"   # Başlık satırı yazısı
C_SUMMARY   = "D6EAF8"   # Özet satırı


def _border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Scraping ────────────────────────────────────────────────────────────────

def fetch_ranking(year: int, subject_code: str) -> list[dict]:
    """
    Belirtilen yıl ve konu koduna ait sıralama sayfasını çeker,
    yalnızca Türkiye üniversitelerini döndürür.
    """
    url = f"{BASE_URL}/ranking/{year}/{subject_code}/"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  ✗ {url} — {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    table = soup.find("table")
    if not table:
        print(f"  ✗ Tablo bulunamadı: {url}")
        return []

    # Sütun başlıkları
    headers_row = table.find("thead")
    if headers_row:
        col_names = [th.get_text(strip=True) for th in headers_row.find_all("th")]
    else:
        first_row = table.find("tr")
        col_names = [td.get_text(strip=True) for td in first_row.find_all(["th", "td"])]

    rows = []
    for tr in table.find("tbody").find_all("tr") if table.find("tbody") else table.find_all("tr")[1:]:
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue

        row_data = {}
        for i, cell in enumerate(cells):
            col = col_names[i] if i < len(col_names) else f"col_{i}"
            # Üniversite adı için linki de al
            link = cell.find("a")
            row_data[col] = cell.get_text(strip=True)
            if link and link.get("href"):
                row_data["_profile_url"] = (
                    link["href"] if link["href"].startswith("http")
                    else BASE_URL + link["href"]
                )

        # Ülke kontrolü — "Country" sütununda "Turkiye" geçenler
        country_val = (
            row_data.get("Country")
            or row_data.get("country")
            or ""
        )
        if COUNTRY.lower() not in country_val.lower():
            continue

        row_data["_year"]    = year
        row_data["_subject"] = subject_code
        rows.append(row_data)

    return rows


def scrape_all(years: list[int] = YEARS, delay: float = 0.4) -> dict:
    """
    Tüm yıl × konu kombinasyonlarını çeker.
    Döndürülen yapı: { subject_name: { year: [rows] } }
    """
    results = {}
    total = len(SUBJECTS) * len(years)
    done  = 0

    for subject_name, code in SUBJECTS.items():
        results[subject_name] = {}
        for year in years:
            done += 1
            print(f"[{done:02d}/{total}] {year} | {subject_name}")
            rows = fetch_ranking(year, code)
            results[subject_name][year] = rows
            print(f"       → {len(rows)} Türk üniversite bulundu")
            time.sleep(delay)

    return results


# ─── JSON çıktısı ────────────────────────────────────────────────────────────

def save_json(data: dict, filename: str) -> None:
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"\n✓ JSON kaydedildi: {filename}")


# ─── Excel çıktısı ───────────────────────────────────────────────────────────

def _write_sheet(wb: Workbook, sheet_name: str, rows_by_year: dict) -> None:
    """
    Bir konu için ayrı bir Excel sayfası yazar.
    Her yılın verisi yan yana bloklar halinde yer alır.
    """
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 karakter

    # Başlık satırı
    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = f"EngiRank — {sheet_name} — Türk Üniversiteleri"
    c.font      = Font(name="Arial", bold=True, size=13, color=C_TITLE_FG)
    c.fill      = PatternFill("solid", start_color=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"].value = f"Oluşturulma tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font  = Font(name="Arial", italic=True, size=9, color="7F8C8D")
    ws.merge_cells("A2:Z2")
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6   # boşluk

    # Temel sütunlar (her yıl bloğu için)
    BASE_COLS = [
        "Rank", "Institution", "Country",
        "Final Score", "Research 28%", "Innovation 25%",
        "SDG 10%", "Internationalization 16%", "Multidisciplinarity 21%",
        "Profile URL",
    ]

    col_offset = 1
    for year in sorted(rows_by_year.keys(), reverse=True):
        rows = rows_by_year[year]
        if not rows:
            continue

        # Yıl başlığı (birleşik hücre)
        end_col = col_offset + len(BASE_COLS) - 1
        ws.merge_cells(
            start_row=4, start_column=col_offset,
            end_row=4,   end_column=end_col
        )
        year_cell = ws.cell(row=4, column=col_offset, value=str(year))
        year_cell.font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FONT)
        year_cell.fill      = PatternFill("solid", start_color=C_HDR_BG)
        year_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 22

        # Sütun başlıkları
        for j, col_name in enumerate(BASE_COLS):
            cell = ws.cell(row=5, column=col_offset + j, value=col_name)
            cell.font      = Font(name="Arial", bold=True, size=10, color=C_HDR_FONT)
            cell.fill      = PatternFill("solid", start_color="2E6DA4")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _border()
        ws.row_dimensions[5].height = 30

        # Veri satırları
        for i, row in enumerate(rows):
            excel_row = 6 + i
            bg = PatternFill("solid", start_color=C_ODD if i % 2 == 0 else C_EVEN)

            # Sıralama değerlerini normalize et — kaynak tablodaki sütun adları değişkenlik gösterebilir
            # Sıralama değerlerini normalize et — kaynak tablodaki sütun adları değişkenlik gösterebilir
            def _get(*keys):
                # 1. Önce tam eşleşme kontrolü
                for k in keys:
                    if k in row:
                        return row[k]
                
                # 2. Tam eşleşme yoksa, büyük/küçük harf duyarsız kısmi (substring) eşleşme kontrolü
                for k in keys:
                    for row_key, row_value in row.items():
                        if isinstance(row_key, str):
                            # Gelen başlıktaki görünmez karakterleri ve satır atlamalarını temizle
                            clean_key = row_key.lower().replace("\n", " ").replace("\xa0", " ")
                            if k.lower() in clean_key:
                                return row_value
                return ""

            values = [
                _get("Rank 2025", "Rank 2024", "Rank 2023", "Rank"),
                _get("Institution"),
                _get("Country"),
                _get("Final Score", "Score"),
                _get("Research"),               # Yüzdeleri kaldırdık, sadece kelimeyi arıyoruz
                _get("Innovation"),
                _get("SDG"),
                _get("Internationalization"),
                _get("Multidisciplinarity"),
                row.get("_profile_url", ""),
            ]

            for j, val in enumerate(values):
                cell = ws.cell(row=excel_row, column=col_offset + j, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.fill      = bg
                cell.border    = _border()
                cell.alignment = Alignment(vertical="center", wrap_text=(j == 1))

                # URL sütunu: tıklanabilir hyperlink
                if j == len(BASE_COLS) - 1 and val:
                    cell.hyperlink = val
                    cell.font = Font(name="Arial", size=10, color="2471A3", underline="single")

            ws.row_dimensions[excel_row].height = 18

        # Özet satırı
        summary_row = 6 + len(rows)
        ws.cell(row=summary_row, column=col_offset,
                value=f"Toplam: {len(rows)} üniversite").font = Font(
            name="Arial", bold=True, size=10, color=C_TITLE_FG)
        for j in range(len(BASE_COLS)):
            c2 = ws.cell(row=summary_row, column=col_offset + j)
            c2.fill   = PatternFill("solid", start_color=C_SUMMARY)
            c2.border = _border()
        ws.row_dimensions[summary_row].height = 18

        # Sütun genişlikleri
        widths = [8, 40, 12, 12, 12, 12, 10, 18, 18, 55]
        for j, w in enumerate(widths):
            col_letter = get_column_letter(col_offset + j)
            ws.column_dimensions[col_letter].width = w

        col_offset = end_col + 2   # yıl blokları arasında 1 boş sütun

    ws.freeze_panes = "A6"


def save_excel(data: dict, filename: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)   # default boş sayfayı sil

    # Özet sayfası
    ws_sum = wb.create_sheet("ÖZET", 0)
    ws_sum.merge_cells("A1:F1")
    c = ws_sum["A1"]
    c.value     = "EngiRank — Türk Üniversiteleri Özet"
    c.font      = Font(name="Arial", bold=True, size=14, color=C_TITLE_FG)
    c.fill      = PatternFill("solid", start_color=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum["A2"].value = f"Oluşturulma tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_sum["A2"].font  = Font(name="Arial", italic=True, size=9, color="7F8C8D")
    ws_sum.merge_cells("A2:F2")

    headers = ["Konu Alanı", "2025 Üniversite Sayısı", "2024 Üniversite Sayısı", "2023 Üniversite Sayısı"]
    for j, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=4, column=j, value=h)
        cell.font      = Font(name="Arial", bold=True, size=11, color=C_HDR_FONT)
        cell.fill      = PatternFill("solid", start_color=C_HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws_sum.row_dimensions[4].height = 28

    for i, (subject_name, rows_by_year) in enumerate(data.items()):
        row = 5 + i
        bg  = PatternFill("solid", start_color=C_ODD if i % 2 == 0 else C_EVEN)
        counts = [rows_by_year.get(y, []) for y in [2025, 2024, 2023]]
        for j, val in enumerate([subject_name] + [len(c) for c in counts], 1):
            cell = ws_sum.cell(row=row, column=j, value=val)
            cell.font   = Font(name="Arial", size=10)
            cell.fill   = bg
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")
        ws_sum.row_dimensions[row].height = 18

    ws_sum.column_dimensions["A"].width = 45
    for col in ["B", "C", "D"]:
        ws_sum.column_dimensions[col].width = 22

    # Her konu için ayrı sayfa
    for subject_name, rows_by_year in data.items():
        print(f"  Excel sayfası yazılıyor: {subject_name}")
        _write_sheet(wb, subject_name, rows_by_year)

    wb.save(filename)
    print(f"\n✓ Excel kaydedildi: {filename}")


# ─── Ana akış ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("  EngiRank Türk Üniversiteleri Scraper")
    print("=" * 60)

    data = scrape_all(years=YEARS, delay=0.4)

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_file  = f"engirank_turkiye_{timestamp}.json"
    excel_file = f"engirank_turkiye_{timestamp}.xlsx"

    save_json(data, json_file)
    save_excel(data, excel_file)

    # Konsol özeti
    print("\n" + "=" * 60)
    print("  ÖZET")
    print("=" * 60)
    for subject, by_year in data.items():
        counts = " | ".join(f"{y}: {len(by_year.get(y,[]))}" for y in YEARS)
        print(f"  {subject:<45} {counts}")