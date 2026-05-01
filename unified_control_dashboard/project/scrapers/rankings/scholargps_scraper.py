"""
ScholarGPS Kusursuz Scraper (Global Metrik + Anlık Tam Kayıt)
======================================================
Bu kod bağlantı kopsa bile çökmez, sizi bekler.
Çektiği ve kalan tüm verileri anında JSON'a kaydeder.
"""

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time
import random
import json
import os
import sys
import io
from pathlib import Path

# Windows terminal encoding fix
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except (AttributeError, io.UnsupportedOperation):
        pass

BASE_DIR = Path(__file__).resolve().parents[2]
OUTPUT_DIR = BASE_DIR / "data" / "raw"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CHECKPOINT_FILE = OUTPUT_DIR / "scholargps_kalan_linkler.json"
COMPLETED_FILE = OUTPUT_DIR / "scholargps_tamamlananlar.json"
INTERACTIVE_MODE = os.getenv("SCHOLARGPS_INTERACTIVE", "0") == "1" or sys.stdin.isatty()
HEADLESS_MODE = os.getenv("SCHOLARGPS_HEADLESS", "0") == "1" # Varsayılan olarak GÖRÜNÜR mod (False)


def _log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def _wait_with_heartbeat(total_seconds: float, reason: str, step_seconds: int = 5) -> None:
    total = max(0, int(total_seconds))
    if total == 0:
        return

    _log(f"Bekleme basladi ({total}s): {reason}")
    remaining = total
    while remaining > 0:
        chunk = step_seconds if remaining > step_seconds else remaining
        time.sleep(chunk)
        remaining -= chunk
        _log(f"  ...islem suruyor, kalan ~{remaining}s")


def _manual_step_or_wait(prompt: str, wait_seconds: int = 20) -> None:
    _log(f"[BİLGİ] {prompt}")
    _log(f"Bu aşamada tarayıcı penceresinde işlem yapmanız gerekebilir. {wait_seconds} saniye bekleniyor...")
    _wait_with_heartbeat(wait_seconds, "kullanıcı müdahalesi veya otomatik bekleme", step_seconds=5)

def get_or_create_links(page):
    # EĞER KALANLAR DOSYASI VARSA YÜKLE
    if CHECKPOINT_FILE.exists():
        with CHECKPOINT_FILE.open("r", encoding="utf-8") as f:
            institutions_list = json.load(f)
            _log(f"📁 Kayıt dosyası bulundu. Kalan universite sayisi: {len(institutions_list)}")
            return institutions_list

    # EĞER DOSYA YOKSA SIFIRDAN 1. AŞAMAYI BAŞLAT
    _log("ASAMA 1: Kurum listesi ve global metrikler toplaniyor...")
    institutions_list = []
    current_page = 1
    
    # Sıfırdan başlıyorsak eski tamamlananlar dosyasını temizleyelim ki karışmasın
    if COMPLETED_FILE.exists():
        COMPLETED_FILE.unlink()
    
    while True:
        url = f"https://scholargps.com/institutional-rankings?year=2025&country=Turkey&p={current_page}"
        _log(f"Liste sayfasi isleniyor: p={current_page}")
        
        try:
            # Cloudflare korumasını aşmak için 'commit' modunda git ve bekle
            page.goto(url, wait_until="commit", timeout=90000)
            
            # Rastgele bir fare hareketi yaparak "insan" olduğumuzu hissettirelim
            page.mouse.move(random.randint(100, 500), random.randint(100, 500))
            
            time.sleep(random.uniform(5.0, 8.0)) 
            
            if "Cloudflare" in page.content() or "Verify you are human" in page.content() or "challenges.cloudflare.com" in page.content():
                _log("⚠️ Cloudflare koruması algılandı! Lütfen pencerede doğrulamayı yapın (Kutucuğa tıklayın).")
                _log("Eğer kabul etmezse lütfen pencere içinden sayfayı yenileyin (F5).")
                # Kullanıcıya müdahale için uzun süre tanı
                for _ in range(12): # Toplam 60 saniye bekleme döngüsü
                    if "table" in page.content() or ("institutional-rankings" in page.url and "p=" in page.url):
                         break
                    time.sleep(5)
                    _log("  ...bekleniyor (doğrulamayı yapmadıysanız lütfen yapın)")
            
            page.wait_for_selector("table", timeout=30000)
        except PlaywrightTimeoutError:
            _log(f"\n[⏳ ZAMAN AŞIMI] Tablo bulunamadı. Korumayı geçememiş olabilirsiniz.")
            _wait_with_heartbeat(10, "Manuel kontrol ve deneme")
            continue 
        except Exception as e:
            _log(f"\n[!] Beklenmeyen bir ağ hatası: {e}")
            _wait_with_heartbeat(15, "Hata sonrasi otomatik bekleme")
            continue

        soup = BeautifulSoup(page.content(), "html.parser")
        table = soup.find("table")
        rows = table.find("tbody").find_all("tr")
        
        if not rows:
            _log("Sayfada kurum bulunamadi, liste tamamlanmis olabilir.")
            break
            
        for row in rows:
            cols = row.find_all("td")
            if len(cols) < 2: continue
            
            link_tag = cols[1].find("a")
            inst_name = cols[1].get_text(strip=True)
            
            institutions_list.append({
                "Ulusal Sıra": cols[0].get_text(strip=True),
                "Üniversite": inst_name,
                "Global Sıralama Metriği": cols[2].get_text(strip=True) if len(cols) > 2 else "N/A",
                "Profil URL": "https://scholargps.com" + link_tag['href'] if link_tag else None
            })
            
        _log(f"  -> {len(rows)} kurum eklendi. Toplam: {len(institutions_list)}")
        
        # Sonraki sayfa kontrolü
        next_btn = soup.find("a", string="Next")
        if not next_btn or current_page >= 10: # Güvenlik için 10 sayfa sınırı (zaten TR için yeterli)
            break
            
        current_page += 1
        _wait_with_heartbeat(random.uniform(2.0, 4.0), "sayfalar arasi nefes", step_seconds=2)

    with CHECKPOINT_FILE.open("w", encoding="utf-8") as f:
        json.dump(institutions_list, f, ensure_ascii=False, indent=2)
        
    _log(f"[OK] Asama 1 tamamlandi. {len(institutions_list)} universitenin linki kaydedildi.")
    return institutions_list

def scrape_scholargps_detailed():
    completed_results = []
    
    # EĞER ÖNCEDEN BAŞARIYLA ÇEKİLMİŞ VERİLER VARSA HAFIZAYA AL
    if COMPLETED_FILE.exists():
        with COMPLETED_FILE.open("r", encoding="utf-8") as f:
            completed_results = json.load(f)
            _log(f"📁 Onceki oturumdan {len(completed_results)} kurum verisi yuklendi.")

    user_data_dir = OUTPUT_DIR / "browser_profile"
    user_data_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        # Dashboard'da görünebilmesi için headless=False
        # launch_persistent_context kullanarak gerçek bir kullanıcı profili simüle ediyoruz
        # 'channel="chrome"' ekleyerek bilgisayardaki gerçek Chrome'u kullanıyoruz
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(user_data_dir),
            channel="chrome", # GERÇEK CHROME KULLANIMI
            headless=HEADLESS_MODE,
            slow_mo=150,
            viewport={'width': 1920, 'height': 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
            timezone_id="Europe/Istanbul",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox"
            ]
        )
        page = context.pages[0] if context.pages else context.new_page()
        
        # Gelişmiş Gizlilik Ayarları (Playwright izlerini siler)
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = { runtime: {} };
            Object.defineProperty(navigator, 'languages', {get: () => ['tr-TR', 'tr', 'en-US', 'en']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        """)

        try:
            institutions_list = get_or_create_links(page)
        except Exception as e:
            _log(f"🛑 Hata: Kurum listesi alınırken bir sorun oluştu: {e}")
            import traceback
            traceback.print_exc()
            context.close()
            return []

        _log("ASAMA 2: Profil sayfalarindan metrikler cekiliyor...")

        # Sayaç için değişkenler
        toplam_kalan = len(institutions_list)
        islenen_sayisi = 0
        stage_started_at = time.time()

        while institutions_list:
            institution = institutions_list[0] 
            target_url = institution['Profil URL']
            elapsed = max(1, int(time.time() - stage_started_at))
            processed = len(completed_results)
            avg = elapsed / processed if processed > 0 else 0
            remaining_items = len(institutions_list)
            eta = int(avg * remaining_items) if avg > 0 else -1
            eta_text = f"~{eta//60}dk {eta%60}sn" if eta >= 0 else "hesaplaniyor"
            _log(
                f"[{islenen_sayisi + 1}/{toplam_kalan}] {institution['Üniversite']} inceleniyor "
                f"(tamamlanan={processed}, kalan={remaining_items}, tahmini kalan={eta_text})"
            )
            
            try:
                page.goto(target_url, wait_until="domcontentloaded", timeout=60000)
            except PlaywrightTimeoutError:
                print(f"\n   [⏳ ZAMAN AŞIMI] {institution['Üniversite']} sayfası 60 saniyede yüklenemedi.")
                _manual_step_or_wait("   Sayfa yüklenemedi, tekrar denenecek.", wait_seconds=20)
                continue
            except Exception as e:
                print(f"\n   [!] Ağ hatası: {e}")
                time.sleep(5)
                continue
                
            try:
                # RATE LIMIT KONTROLÜ
                page_text = page.content().lower()
                if "rate limit exceeded" in page_text or "too many requests" in page_text:
                    _log("[RATE LIMIT] Gecici engel algilandi. 300s beklenip ayni kurum tekrar denenecek.")
                    _wait_with_heartbeat(300, "rate limit soguma", step_seconds=15)
                    _log("[RATE LIMIT] Bekleme bitti, ayni kurum yeniden deneniyor.")
                    page.reload() 
                    _wait_with_heartbeat(5, "sayfa yenileme sonrasi bekleme", step_seconds=1)
                    continue
                
                try:
                    page.wait_for_selector('.box_left_column_value', timeout=15000) 
                except:
                    print(f"\n   [🚨 DİKKAT] {institution['Üniversite']} sayfasında CAPTCHA veya engel tespit edildi!")
                    _manual_step_or_wait("   CAPTCHA/engel tespit edildi.", wait_seconds=30)
                
                # İnsansı Bekleme
                bekleme = random.uniform(8.0, 15.0)
                _wait_with_heartbeat(bekleme, "insansi bekleme", step_seconds=3)
                
                sub_soup = BeautifulSoup(page.content(), 'html.parser')
                metric_values = sub_soup.find_all('div', class_='box_left_column_value')
                
                if len(metric_values) >= 3:
                    institution["Yayın Sayısı"] = metric_values[0].get_text(strip=True)
                    institution["Atıf Sayısı"] = metric_values[1].get_text(strip=True)
                    institution["Akademisyen Sayısı"] = metric_values[2].get_text(strip=True)
                else:
                    _log(f"[UYARI] {institution['Üniversite']} icin metrik kutulari bulunamadi.")

                # Başarılı kurumu kaydedilenlere ekle, kalanlardan çıkar
                completed_results.append(institution)
                institutions_list.pop(0)
                islenen_sayisi += 1 
                
                # ANINDA KAYIT SİSTEMİ (Elektrik gitse bile veri kaybolmaz)
                with CHECKPOINT_FILE.open("w", encoding="utf-8") as f:
                    json.dump(institutions_list, f, ensure_ascii=False, indent=2)
                with COMPLETED_FILE.open("w", encoding="utf-8") as f:
                    json.dump(completed_results, f, ensure_ascii=False, indent=2)
                _log(
                    f"Kaydedildi: {institution['Üniversite']} | kalan={len(institutions_list)} "
                    f"| tamamlanan={len(completed_results)}"
                )

            except Exception as e:
                _log(f"[HATA] {institution['Üniversite']} islenirken hata: {e}")
                _wait_with_heartbeat(10, "hata sonrasi bekleme", step_seconds=2)
                failed_inst = institutions_list.pop(0)
                institutions_list.append(failed_inst)

        context.close()
        
        # Her şey bittiğinde JSON dosyalarını temizle
        if CHECKPOINT_FILE.exists():
            CHECKPOINT_FILE.unlink()
        if COMPLETED_FILE.exists():
            COMPLETED_FILE.unlink()
            
        return completed_results

def save_to_excel_detailed(data, filename):
    if not data:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "TR Üniversiteleri Detaylı"

    # Başlık Alanı
    ws.merge_cells("A1:G1") # Sütun sayısı 7'ye çıktığı için G1'e uzatıldı
    title_cell = ws["A1"]
    title_cell.value = "ScholarGPS - Türk Üniversiteleri Detaylı Alt Metrikleri"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="8E44AD")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Yeni Sütun Eklendi ("Global Sıralama Metriği")
    headers = ["Ulusal Sıra", "Üniversite Adı", "Global Sıralama Metriği", "Yayın Sayısı", "Atıf Sayısı", "Akademisyen Sayısı", "Profil URL"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2C3E50")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"), top=Side(style="thin", color="BDC3C7"), bottom=Side(style="thin", color="BDC3C7"))

    for i, row_data in enumerate(data):
        row_num = 4 + i
        bg_color = "ECF0F1" if i % 2 == 0 else "FFFFFF"
        
        values = [
            row_data.get("Ulusal Sıra"),
            row_data.get("Üniversite"),
            row_data.get("Global Sıralama Metriği", "N/A"), # Eski jsonlarla çakışmasın diye .get kullanıldı
            row_data.get("Yayın Sayısı"),
            row_data.get("Atıf Sayısı"),
            row_data.get("Akademisyen Sayısı"),
            row_data.get("Profil URL")
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg_color)
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx in [2,3,7] else "center")
            cell.border = Border(left=Side(style="thin", color="BDC3C7"), right=Side(style="thin", color="BDC3C7"), top=Side(style="thin", color="BDC3C7"), bottom=Side(style="thin", color="BDC3C7"))

            if col_idx == 7 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="2980B9", underline="single")

    # Genişlikler Güncellendi
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 60

    wb.save(str(filename))
    _log(f"[OK] Excel dosyasi olusturuldu: {filename}")

if __name__ == "__main__":
    try:
        results = scrape_scholargps_detailed()
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    if results:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = OUTPUT_DIR / f"scholargps_final_rapor_{timestamp}.xlsx"
        save_to_excel_detailed(results, excel_filename)
        _log(f"Toplam tamamlanan kurum: {len(results)}")
    else:
        _log("Sonuc uretilmedi. Mevcut checkpoint dosyalarini kontrol et.")