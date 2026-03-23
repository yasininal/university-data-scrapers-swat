from __future__ import annotations

import csv
import io
import json
import sqlite3
import subprocess
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook, Workbook
except Exception:  # pragma: no cover
    load_workbook = None
    Workbook = None

try:
    from unified_control_dashboard.unified_control.config import (
        REPO_ROOT,
        PYTHON_EXECUTABLE,
        HISTORY_FILE,
        HISTORY_LIMIT,
        RUN_TIMEOUT_SECONDS,
    )
    from unified_control_dashboard.unified_control.models import RunResult
    from unified_control_dashboard.unified_control.registry import build_jobs
    from unified_control_dashboard.unified_control.runner import run_job
    from unified_control_dashboard.unified_control.history import append_history, load_history
    from unified_control_dashboard.unified_control.orchestration import PIPELINES, ordered_jobs_with_dependencies
    from unified_control_dashboard.unified_control.health import check_modules, check_paths, port_in_use
except ModuleNotFoundError:
    from unified_control.config import (
        REPO_ROOT,
        PYTHON_EXECUTABLE,
        HISTORY_FILE,
        HISTORY_LIMIT,
        RUN_TIMEOUT_SECONDS,
    )
    from unified_control.models import RunResult
    from unified_control.registry import build_jobs
    from unified_control.runner import run_job
    from unified_control.history import append_history, load_history
    from unified_control.orchestration import PIPELINES, ordered_jobs_with_dependencies
    from unified_control.health import check_modules, check_paths, port_in_use


BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIST_DIR = BASE_DIR / "frontend" / "dist"
FRONTEND_ASSETS_DIR = FRONTEND_DIST_DIR / "assets"

app = FastAPI(title="Unified Control Dashboard API", version="1.0.0")
if FRONTEND_ASSETS_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_ASSETS_DIR)), name="assets")

RUNS_LOCK = threading.Lock()
RUNS: dict[str, dict[str, Any]] = {}
MAX_RUNS = 100
MAX_LOG_LINES = 250


def _fmt_ts(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def _safe_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("\n", " ").strip()


def _collect_output_files(expected_outputs: list[str]) -> list[Path]:
    files: list[Path] = []
    for pattern in expected_outputs:
        matches = sorted(REPO_ROOT.glob(pattern), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
        if matches:
            files.append(matches[0])
        else:
            direct = REPO_ROOT / pattern
            if direct.exists():
                files.append(direct)
    unique: dict[str, Path] = {}
    for item in files:
        unique[str(item.resolve())] = item
    return sorted(unique.values(), key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)


def _json_preview(path: Path) -> dict[str, Any] | None:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None

    if isinstance(payload, list) and payload and isinstance(payload[0], dict):
        rows = payload[:15]
        columns = list(rows[0].keys())[:10]
        table_rows = [[_safe_text(row.get(col, "")) for col in columns] for row in rows]
        return {
            "title": f"Veri Onizleme ({path.name})",
            "columns": columns,
            "rows": table_rows,
            "total_rows": len(payload),
        }

    if isinstance(payload, dict):
        list_key = None
        for key, value in payload.items():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                list_key = key
                break
        if list_key is not None:
            rows = payload[list_key][:15]
            columns = list(rows[0].keys())[:10]
            table_rows = [[_safe_text(row.get(col, "")) for col in columns] for row in rows]
            return {
                "title": f"Veri Onizleme ({path.name} / {list_key})",
                "columns": columns,
                "rows": table_rows,
                "total_rows": len(payload[list_key]),
            }

        kv_rows = [[str(k), _safe_text(v)] for k, v in list(payload.items())[:20]]
        return {
            "title": f"Anahtar Deger Onizleme ({path.name})",
            "columns": ["Alan", "Deger"],
            "rows": kv_rows,
            "total_rows": len(payload),
        }

    return None


def _csv_preview(path: Path) -> dict[str, Any] | None:
    try:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.DictReader(handle)
            columns = reader.fieldnames or []
            columns = columns[:10]
            rows = []
            total = 0
            for item in reader:
                total += 1
                if len(rows) < 15:
                    rows.append([_safe_text(item.get(col, "")) for col in columns])
    except Exception:
        return None

    if not columns:
        return None
    return {
        "title": f"Veri Onizleme ({path.name})",
        "columns": columns,
        "rows": rows,
        "total_rows": total,
    }


def _db_preview(path: Path) -> dict[str, Any] | None:
    try:
        conn = sqlite3.connect(str(path))
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [r[0] for r in cur.fetchall()]
        rows = []
        for table in tables[:20]:
            cur.execute(f'SELECT COUNT(*) FROM "{table}"')
            count = cur.fetchone()[0]
            rows.append([table, str(count)])
        conn.close()
    except Exception:
        return None

    return {
        "title": f"Veritabani Ozeti ({path.name})",
        "columns": ["Tablo", "Kayit Sayisi"],
        "rows": rows,
        "total_rows": len(rows),
    }


def _build_data_preview(files: list[Path]) -> dict[str, Any] | None:
    for file_path in files:
        suffix = file_path.suffix.lower()
        if suffix == ".json":
            preview = _json_preview(file_path)
            if preview:
                return preview
        if suffix == ".csv":
            preview = _csv_preview(file_path)
            if preview:
                return preview
        if suffix in {".db", ".sqlite", ".sqlite3"}:
            preview = _db_preview(file_path)
            if preview:
                return preview
    return None


def _read_json_table(path: Path) -> tuple[list[str], list[list[str]], str]:
    payload = json.loads(path.read_text(encoding="utf-8"))

    if isinstance(payload, list) and payload and isinstance(payload[0], dict):
        columns = list(payload[0].keys())
        rows = [[_safe_text(item.get(col, "")) for col in columns] for item in payload]
        return columns, rows, path.name

    if isinstance(payload, dict):
        for key, value in payload.items():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                columns = list(value[0].keys())
                rows = [[_safe_text(item.get(col, "")) for col in columns] for item in value]
                return columns, rows, f"{path.name}:{key}"
        columns = ["key", "value"]
        rows = [[str(k), _safe_text(v)] for k, v in payload.items()]
        return columns, rows, path.name

    raise ValueError("JSON veri formati tabloya cevrilemedi")


def _read_csv_table(path: Path) -> tuple[list[str], list[list[str]], str]:
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = reader.fieldnames or []
        rows = [[_safe_text(item.get(col, "")) for col in columns] for item in reader]
    if not columns:
        raise ValueError("CSV kolonlari bulunamadi")
    return columns, rows, path.name


def _read_sqlite_table(path: Path) -> tuple[list[str], list[list[str]], str]:
    conn = sqlite3.connect(str(path))
    cur = conn.cursor()

    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [r[0] for r in cur.fetchall()]
    if not tables:
        conn.close()
        raise ValueError("Veritabaninda tablo bulunamadi")

    selected = "grants" if "grants" in tables else tables[0]
    cur.execute(f'SELECT * FROM "{selected}"')
    data = cur.fetchall()
    columns = [item[0] for item in (cur.description or [])]
    conn.close()

    rows = [[_safe_text(cell) for cell in row] for row in data]
    return columns, rows, f"{path.name}:{selected}"


def _read_xlsx_table(path: Path) -> tuple[list[str], list[list[str]], str]:
    if pd is not None:
        try:
            df = pd.read_excel(path)
            columns = [str(col).strip() or f"column_{idx + 1}" for idx, col in enumerate(df.columns.tolist())]
            rows = [
                [_safe_text(value) for value in row]
                for row in df.where(pd.notna(df), None).values.tolist()
            ]
            return columns, rows, path.name
        except Exception:
            pass

    if load_workbook is None:
        raise ValueError("openpyxl kurulu degil")

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not all_rows:
        wb.close()
        raise ValueError("Excel dosyasi bos")

    # Try to locate a meaningful header row for formatted workbooks.
    scan_limit = min(len(all_rows), 30)
    best_index = 0
    best_score = -1
    for idx in range(scan_limit):
        row = all_rows[idx]
        non_empty = sum(1 for cell in row if _safe_text(cell))
        if non_empty > best_score:
            best_score = non_empty
            best_index = idx

    header_row = all_rows[best_index]
    last_non_empty = -1
    for idx, cell in enumerate(header_row):
        if _safe_text(cell):
            last_non_empty = idx

    if last_non_empty < 0:
        wb.close()
        raise ValueError("Excel baslik satiri bulunamadi")

    width = last_non_empty + 1
    columns: list[str] = []
    for idx in range(width):
        value = _safe_text(header_row[idx])
        columns.append(value or f"column_{idx + 1}")

    table_rows: list[list[str]] = []
    for row in all_rows[best_index + 1 :]:
        normalized = [_safe_text(row[idx] if idx < len(row) else "") for idx in range(width)]
        if any(cell for cell in normalized):
            table_rows.append(normalized)

    wb.close()
    return columns, table_rows, f"{path.name}:{ws.title}"


def _read_full_table(path: Path) -> tuple[list[str], list[list[str]], str]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _read_json_table(path)
    if suffix == ".csv":
        return _read_csv_table(path)
    if suffix in {".db", ".sqlite", ".sqlite3"}:
        return _read_sqlite_table(path)
    if suffix == ".xlsx":
        return _read_xlsx_table(path)
    raise ValueError(f"Desteklenmeyen dosya uzantisi: {suffix}")


def _latest_file_for_job(job) -> Path:
    files = _collect_output_files(job.expected_outputs)
    if not files:
        raise ValueError("Bu scraper icin veri dosyasi bulunamadi")
    return files[0]


def _best_data_file_for_job(job) -> Path:
    files = _collect_output_files(job.expected_outputs)
    if not files:
        raise ValueError("Bu scraper icin veri dosyasi bulunamadi")

    best_file = files[0]
    best_score = (-1, -1)
    for file_path in files:
        try:
            columns, rows, _ = _read_full_table(file_path)
        except Exception:
            continue

        # Prioritize non-empty datasets; then richer column count.
        score = (len(rows), len(columns))
        if score > best_score:
            best_score = score
            best_file = file_path

    return best_file


def _build_excel_stream(columns: list[str], rows: list[list[str]]) -> io.BytesIO:
    if Workbook is None:
        raise ValueError("Excel olusturmak icin openpyxl gerekli")

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(columns)
    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_result_table(job, result: dict[str, Any]) -> dict[str, Any]:
    files = _collect_output_files(job.expected_outputs)
    file_rows = []
    for file_path in files:
        stats = file_path.stat()
        file_rows.append([
            str(file_path.relative_to(REPO_ROOT)),
            _fmt_ts(stats.st_mtime),
            str(round(stats.st_size / 1024, 1)),
        ])

    preview = _build_data_preview(files)
    return {
        "summary": {
            "columns": ["Alan", "Deger"],
            "rows": [
                ["Script", result.get("job_name", "")],
                ["Durum", "Basarili" if result.get("success") else "Hatali"],
                ["Calisma Suresi (sn)", str(result.get("duration_seconds", ""))],
                ["Cikis Kodu", str(result.get("exit_code", ""))],
                ["Baslangic", result.get("started_at", "")],
                ["Bitis", result.get("finished_at", "")],
            ],
        },
        "files": {
            "title": "Uretilen Dosyalar",
            "columns": ["Dosya", "Guncelleme", "Boyut (KB)"],
            "rows": file_rows,
        },
        "preview": preview,
    }


def _enrich_result(job, result: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(result)
    enriched["result_table"] = _build_result_table(job, enriched)
    return enriched


def _job_to_dict(job) -> dict[str, Any]:
    return {
        "id": job.id,
        "name": job.name,
        "category": job.category,
        "description": job.description,
        "command": job.command,
        "cwd": str(job.cwd),
        "expected_outputs": job.expected_outputs,
        "dependencies": job.dependencies,
    }


def _trim_runs() -> None:
    if len(RUNS) <= MAX_RUNS:
        return
    sorted_ids = sorted(RUNS, key=lambda rid: RUNS[rid].get("started_at", ""))
    for rid in sorted_ids[: max(0, len(sorted_ids) - MAX_RUNS)]:
        RUNS.pop(rid, None)


def _update_run(run_id: str, **kwargs: Any) -> None:
    with RUNS_LOCK:
        state = RUNS.get(run_id)
        if state is None:
            return
        state.update(kwargs)


def _append_run_log(run_id: str, line: str) -> None:
    clean = line.rstrip("\n")
    if not clean:
        return
    with RUNS_LOCK:
        state = RUNS.get(run_id)
        if state is None:
            return
        logs = state.get("log_lines", [])
        logs.append(clean)
        if len(logs) > MAX_LOG_LINES:
            logs[:] = logs[-MAX_LOG_LINES:]
        state["log_lines"] = logs
        state["last_log_line"] = clean
        state["updated_at"] = RunResult.now_iso()


def _run_job_async(run_id: str, job, timeout_seconds: int) -> None:
    started = datetime.utcnow()
    t0 = time.time()
    exit_code = 1
    timed_out = False
    full_logs: list[str] = []

    try:
        process = subprocess.Popen(
            job.command,
            cwd=str(job.cwd),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
    except Exception as exc:
        _update_run(
            run_id,
            status="failed",
            success=False,
            exit_code=1,
            finished_at=RunResult.now_iso(),
            duration_seconds=round(time.time() - t0, 2),
            error=f"Runner baslatilamadi: {exc}",
        )
        return

    def _reader() -> None:
        if process.stdout is None:
            return
        for line in process.stdout:
            full_logs.append(line)
            _append_run_log(run_id, line)

    reader_thread = threading.Thread(target=_reader, daemon=True)
    reader_thread.start()

    while reader_thread.is_alive():
        elapsed = time.time() - t0
        _update_run(run_id, duration_seconds=round(elapsed, 2), updated_at=RunResult.now_iso())
        if elapsed > timeout_seconds:
            timed_out = True
            process.kill()
            break
        reader_thread.join(timeout=1.0)

    reader_thread.join(timeout=3.0)
    exit_code = process.poll() if process.poll() is not None else 1

    finished = datetime.utcnow()
    stdout = "".join(full_logs)
    stderr = ""
    if timed_out:
        exit_code = 124
        stderr = "Process timed out."

    success = (exit_code == 0) and not timed_out
    result = RunResult(
        job_id=job.id,
        job_name=job.name,
        success=success,
        exit_code=exit_code,
        started_at=started.replace(microsecond=0).isoformat() + "Z",
        finished_at=finished.replace(microsecond=0).isoformat() + "Z",
        duration_seconds=round(time.time() - t0, 2),
        stdout=stdout[-12000:],
        stderr=stderr[-12000:],
    )

    append_history(HISTORY_FILE, result, max_items=HISTORY_LIMIT)
    enriched = _enrich_result(job, result.to_dict())
    _update_run(
        run_id,
        status="completed" if success else "failed",
        success=success,
        exit_code=exit_code,
        finished_at=result.finished_at,
        duration_seconds=result.duration_seconds,
        result=enriched,
        updated_at=RunResult.now_iso(),
        error=stderr or None,
    )


def _start_async_run(job, timeout_seconds: int) -> str:
    run_id = str(uuid.uuid4())
    state = {
        "run_id": run_id,
        "job_id": job.id,
        "job_name": job.name,
        "status": "running",
        "success": None,
        "exit_code": None,
        "started_at": RunResult.now_iso(),
        "finished_at": None,
        "duration_seconds": 0.0,
        "log_lines": [],
        "last_log_line": "Scraper baslatildi...",
        "updated_at": RunResult.now_iso(),
        "error": None,
        "result": None,
    }
    with RUNS_LOCK:
        RUNS[run_id] = state
        _trim_runs()

    worker = threading.Thread(target=_run_job_async, args=(run_id, job, timeout_seconds), daemon=True)
    worker.start()
    return run_id


def _run_and_store(job, timeout_s: int) -> dict[str, Any]:
    result = run_job(job, timeout_s)
    append_history(HISTORY_FILE, result, max_items=HISTORY_LIMIT)
    return _enrich_result(job, result.to_dict())


@app.get("/")
def root() -> FileResponse:
    vue_index = FRONTEND_DIST_DIR / "index.html"
    if vue_index.exists():
        return FileResponse(vue_index)
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/jobs")
def get_jobs() -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    return {
        "jobs": [_job_to_dict(j) for j in jobs.values()],
        "pipelines": PIPELINES,
        "timeout_default": RUN_TIMEOUT_SECONDS,
        "repo_root": str(REPO_ROOT),
        "python_executable": PYTHON_EXECUTABLE,
    }


@app.post("/api/run/{job_id}")
def run_single_job(job_id: str, timeout_seconds: int = RUN_TIMEOUT_SECONDS) -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Unknown job id: {job_id}")
    return _run_and_store(job, timeout_seconds)


@app.post("/api/run-async/{job_id}")
def run_single_job_async(job_id: str, timeout_seconds: int = RUN_TIMEOUT_SECONDS) -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Unknown job id: {job_id}")

    run_id = _start_async_run(job, timeout_seconds)
    return {"run_id": run_id, "job_id": job.id, "job_name": job.name, "status": "running"}


@app.get("/api/run-status/{run_id}")
def get_run_status(run_id: str) -> dict[str, Any]:
    with RUNS_LOCK:
        state = RUNS.get(run_id)
        if state is None:
            raise HTTPException(status_code=404, detail=f"Unknown run id: {run_id}")
        return dict(state)


@app.post("/api/run-pipeline/{pipeline_name}")
def run_pipeline(pipeline_name: str, timeout_seconds: int = RUN_TIMEOUT_SECONDS) -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    if pipeline_name not in PIPELINES:
        raise HTTPException(status_code=404, detail=f"Unknown pipeline: {pipeline_name}")

    requested_ids = PIPELINES[pipeline_name]
    try:
        ordered = ordered_jobs_with_dependencies(jobs, requested_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    results = []
    for job in ordered:
        item = _run_and_store(job, timeout_seconds)
        results.append(item)
        if not item.get("success"):
            break

    return {
        "pipeline": pipeline_name,
        "requested": requested_ids,
        "executed": [r["job_id"] for r in results],
        "results": results,
    }


@app.get("/api/history")
def get_history(limit: int = 50) -> dict[str, Any]:
    limit = max(1, min(limit, HISTORY_LIMIT))
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    raw_items = load_history(HISTORY_FILE)[:limit]
    items = []
    for item in raw_items:
        job_id = item.get("job_id", "")
        job = jobs.get(job_id)
        if job is None:
            items.append(item)
            continue
        items.append(_enrich_result(job, item))
    return {"items": items}


@app.get("/api/health")
def get_health() -> dict[str, Any]:
    required_paths = [
        "scrapers/rankings",
        "hibe/backend/run_scrapers.py",
        "Surdurulebilirlik_Projesi/itu_surdurulebilirlik.py",
    ]
    modules = [
        "requests",
        "bs4",
        "pandas",
        "openpyxl",
        "sqlalchemy",
        "fastapi",
        "uvicorn",
    ]

    path_status = check_paths(REPO_ROOT, required_paths)
    module_status = check_modules(modules)

    return {
        "repo_root": str(REPO_ROOT),
        "python_executable": PYTHON_EXECUTABLE,
        "path_status": path_status,
        "module_status": module_status,
        "ports": {
            "8000_in_use": port_in_use(8000),
            "8501_in_use": port_in_use(8501),
            "8080_in_use": port_in_use(8080),
        },
    }


@app.get("/api/data/{job_id}")
def get_job_full_data(job_id: str) -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Unknown job id: {job_id}")

    try:
        source_file = _best_data_file_for_job(job)
        columns, rows, source_name = _read_full_table(source_file)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "job_id": job.id,
        "job_name": job.name,
        "source_file": str(source_file.relative_to(REPO_ROOT)),
        "source_name": source_name,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
    }


@app.get("/api/data/{job_id}/all")
def get_job_all_data(job_id: str) -> dict[str, Any]:
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Unknown job id: {job_id}")

    source_files = _collect_output_files(job.expected_outputs)
    if not source_files:
        raise HTTPException(status_code=400, detail="Bu scraper icin veri dosyasi bulunamadi")

    tables: list[dict[str, Any]] = []
    for source_file in source_files:
        try:
            columns, rows, source_name = _read_full_table(source_file)
        except Exception:
            continue

        tables.append(
            {
                "source_file": str(source_file.relative_to(REPO_ROOT)),
                "source_name": source_name,
                "columns": columns,
                "rows": rows,
                "row_count": len(rows),
            }
        )

    if not tables:
        raise HTTPException(status_code=400, detail="Veri dosyalari okunamadi")

    tables.sort(key=lambda item: (item["row_count"], len(item["columns"])), reverse=True)

    return {
        "job_id": job.id,
        "job_name": job.name,
        "table_count": len(tables),
        "tables": tables,
    }


@app.get("/api/data/{job_id}/download.xlsx")
def download_job_data_excel(job_id: str):
    jobs = build_jobs(REPO_ROOT, PYTHON_EXECUTABLE)
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Unknown job id: {job_id}")

    try:
        source_file = _best_data_file_for_job(job)
        if source_file.suffix.lower() == ".xlsx":
            return FileResponse(
                source_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"{job.id}.xlsx",
            )

        columns, rows, _ = _read_full_table(source_file)
        stream = _build_excel_stream(columns, rows)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{job.id}.xlsx"'},
    )


@app.get("/{full_path:path}")
def spa_fallback(full_path: str):
    if full_path.startswith("api/") or full_path.startswith("assets/"):
        raise HTTPException(status_code=404, detail="Not found")

    vue_index = FRONTEND_DIST_DIR / "index.html"
    if vue_index.exists():
        return FileResponse(vue_index)
    raise HTTPException(status_code=503, detail="Frontend build bulunamadi. frontend/dist olusturulmeli.")
